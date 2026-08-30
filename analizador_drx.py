import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


def _a_numero(valor):
    texto = str(valor).strip().replace(",", ".")
    texto = re.sub(r"[^0-9eE+\-.]", "", texto)
    if texto in ("", ".", "-", "+"):
        raise ValueError("valor no numerico")
    return float(texto)


def _tokens_numericos(linea):
    patron = r"[-+]?\d+(?:[.,]\d+)?(?:[eE][-+]?\d+)?"
    return [float(valor.replace(",", ".")) for valor in re.findall(patron, linea)]


def _intentar_parseo_tabular(ruta_archivo):
    """
    Prueba varias combinaciones de separador/decimal (tab, coma, punto y coma,
    espacios, autodeteccion) para leer un archivo de datos experimentales
    con distintos formatos de exportacion.
    """
    intentos = [
        {"sep": "\t", "decimal": ","},
        {"sep": "\t", "decimal": "."},
        {"sep": ",", "decimal": "."},
        {"sep": ";", "decimal": ","},
        {"sep": ";", "decimal": "."},
        {"sep": r"\s+", "decimal": "."},
        {"sep": None, "decimal": "."},
    ]

    for opciones in intentos:
        try:
            tabla = pd.read_csv(ruta_archivo, engine="python", **opciones)
        except Exception:
            continue

        if tabla.shape[1] < 2:
            continue

        columnas = {str(col).strip().lower(): col for col in tabla.columns}
        theta_col = next((
            original for nombre, original in columnas.items()
            if "2th" in nombre or "pos" in nombre or "theta" in nombre or "angulo" in nombre
        ), None)
        intensidad_col = next((
            original for nombre, original in columnas.items()
            if "iobs" in nombre or "intens" in nombre or "cts" in nombre or "counts" in nombre
        ), None)

        if theta_col is None or intensidad_col is None:
            candidatas = tabla.apply(lambda col: pd.to_numeric(col, errors="coerce")).dropna(axis=1, how="all")
            numericas = [c for c in candidatas.columns if candidatas[c].notna().sum() > 3]
            if len(numericas) >= 2:
                theta_col, intensidad_col = numericas[0], numericas[1]
            else:
                continue

        df = tabla[[theta_col, intensidad_col]].copy()
        df.columns = ["2Theta", "Iobs"]
        df["2Theta"] = pd.to_numeric(df["2Theta"], errors="coerce")
        df["Iobs"] = pd.to_numeric(df["Iobs"], errors="coerce")
        df = df.dropna().sort_values("2Theta").drop_duplicates("2Theta").reset_index(drop=True)

        if len(df) >= 3:
            return df

    return None


def CargarDatos(ruta_archivo=None):
    """
    Carga un archivo experimental DRX en (casi) cualquier formato de texto:
    intenta primero varias combinaciones de separador/decimal tabulares y,
    si ninguna funciona, cae a un modo de extraccion por expresiones
    regulares que busca pares numericos linea por linea. Devuelve un
    DataFrame con columnas 2Theta e Iobs.
    """
    if ruta_archivo is None:
        ruta_archivo = filedialog.askopenfilename(
            title="Selecciona un archivo DRX",
            filetypes=[("Archivos de texto", "*.txt *.xy *.dat *.csv"), ("Todos", "*.*")],
        )
    if not ruta_archivo:
        return None

    df = _intentar_parseo_tabular(ruta_archivo)
    if df is not None:
        return df

    filas = []
    with open(ruta_archivo, "r", encoding="utf-8", errors="ignore") as archivo:
        for linea in archivo:
            numeros = _tokens_numericos(linea)
            if len(numeros) >= 2:
                filas.append((numeros[0], numeros[1]))

    if not filas:
        raise ValueError("No se encontraron pares numericos 2Theta/Iobs en el archivo.")

    df = pd.DataFrame(filas, columns=["2Theta", "Iobs"])
    df = df.apply(pd.to_numeric, errors="coerce").dropna()
    df = df.sort_values("2Theta").drop_duplicates("2Theta").reset_index(drop=True)
    if df.empty:
        raise ValueError("El archivo no contiene datos experimentales validos.")
    return df


def CargarReferencia(ruta_archivo=None):
    """
    Carga Referencia.txt. Acepta formatos como:
    18.73 (111), 18,73 111, o columnas con encabezado 2Theta/Indice.
    """
    if ruta_archivo is None:
        ruta_archivo = filedialog.askopenfilename(
            title="Selecciona el archivo de referencia",
            filetypes=[("Archivos de texto", "*.txt *.csv *.dat"), ("Todos", "*.*")],
        )
    if not ruta_archivo:
        return None

    try:
        tabla = pd.read_csv(ruta_archivo, sep=r"\s+|\t|;", engine="python", decimal=".")
        columnas = {str(col).strip().lower(): col for col in tabla.columns}
        theta_col = next((
            original for nombre, original in columnas.items()
            if "2theta" in nombre or "2th" in nombre or nombre == "theta"
        ), None)
        indice_col = next((
            original for nombre, original in columnas.items()
            if "indice" in nombre or "hkl" in nombre or "plano" in nombre
        ), None)
        if theta_col is not None and indice_col is not None:
            df = tabla[[theta_col, indice_col]].copy()
            df.columns = ["2Theta", "Indice"]
            df["2Theta"] = pd.to_numeric(df["2Theta"], errors="coerce")
            df["Indice"] = df["Indice"].astype(str).str.strip()
            df = df.dropna().sort_values("2Theta").reset_index(drop=True)
            if not df.empty:
                return df
    except Exception:
        pass

    filas = []
    with open(ruta_archivo, "r", encoding="utf-8", errors="ignore") as archivo:
        for linea in archivo:
            texto = linea.strip()
            if not texto:
                continue
            match_angulo = re.search(r"[-+]?\d+(?:[.,]\d+)?", texto)
            if match_angulo is None:
                continue
            try:
                angulo = _a_numero(match_angulo.group(0))
            except ValueError:
                continue

            hkl_match = re.search(r"\(?\s*\d+\s+\d+\s+\d+\s*\)?|\(\s*\d+\s*\d+\s*\d+\s*\)", texto)
            if hkl_match:
                indice = hkl_match.group(0).replace(" ", "")
                if not indice.startswith("("):
                    indice = f"({indice})"
            else:
                resto = texto[match_angulo.end():].strip(" ;,\t")
                indice = resto.split()[0] if resto else "Referencia"
            filas.append((angulo, indice))

    if not filas:
        raise ValueError("No se encontraron picos teoricos en el archivo de referencia.")

    return pd.DataFrame(filas, columns=["2Theta", "Indice"]).sort_values("2Theta").reset_index(drop=True)


def CargarCIF(ruta_archivo=None):
    """
    Lector basico de archivos .cif: extrae parametros de celda, grupo
    espacial (si esta presente) y la tabla de posiciones atomicas.
    No depende de librerias externas (gemmi/pymatgen); si el .cif tiene un
    formato muy particular puede no capturar todo, pero cubre el caso
    estandar exportado por la mayoria de bases de datos (COD, ICSD, etc).
    """
    if ruta_archivo is None:
        ruta_archivo = filedialog.askopenfilename(
            title="Selecciona un archivo CIF",
            filetypes=[("CIF", "*.cif"), ("Todos", "*.*")],
        )
    if not ruta_archivo:
        return None

    with open(ruta_archivo, "r", encoding="utf-8", errors="ignore") as archivo:
        contenido = archivo.read()

    def _extraer_valor(etiqueta):
        match = re.search(rf"{etiqueta}\s+([^\s#]+)", contenido)
        if not match:
            return None
        crudo = re.sub(r"\(\d+\)", "", match.group(1))  # quita incertidumbre tipo 5.123(4)
        try:
            return float(crudo)
        except ValueError:
            return crudo.strip("'\"")

    celda = {
        "a": _extraer_valor(r"_cell_length_a"),
        "b": _extraer_valor(r"_cell_length_b"),
        "c": _extraer_valor(r"_cell_length_c"),
        "alpha": _extraer_valor(r"_cell_angle_alpha"),
        "beta": _extraer_valor(r"_cell_angle_beta"),
        "gamma": _extraer_valor(r"_cell_angle_gamma"),
        "volumen": _extraer_valor(r"_cell_volume"),
    }

    grupo_match = re.search(
        r"_(?:symmetry_space_group_name_H-M|space_group_name_H-M_alt)\s+['\"]?([^'\"\n]+)",
        contenido,
    )
    grupo_espacial = grupo_match.group(1).strip() if grupo_match else None

    atomos = _extraer_loop_atomos(contenido)

    return {
        "celda": celda,
        "grupo_espacial": grupo_espacial,
        "atomos": atomos,
        "ruta": ruta_archivo,
    }


def _extraer_loop_atomos(contenido):
    """Busca el bloque 'loop_' que contiene _atom_site_label y arma un DataFrame."""
    lineas = contenido.splitlines()
    for i, linea in enumerate(lineas):
        if linea.strip() != "loop_":
            continue

        encabezados = []
        j = i + 1
        while j < len(lineas) and lineas[j].strip().startswith("_"):
            encabezados.append(lineas[j].strip())
            j += 1

        if not any("_atom_site_label" in encabezado for encabezado in encabezados):
            continue

        filas = []
        while j < len(lineas):
            fila = lineas[j].strip()
            if not fila or fila.startswith("_") or fila.startswith("loop_") or fila.startswith("#"):
                break
            valores = fila.split()
            if len(valores) >= len(encabezados):
                filas.append(valores[: len(encabezados)])
            j += 1

        if filas:
            nombres_col = [
                encabezado.split("_atom_site_")[-1] if "_atom_site_" in encabezado else encabezado
                for encabezado in encabezados
            ]
            return pd.DataFrame(filas, columns=nombres_col)

    return pd.DataFrame()


def RemoverBackground(df, ventana=31, percentil=10, suavizado=9):
    """
    Estima la linea base con un percentil movil y la resta de la intensidad.
    No depende de scipy ni peakutils.
    """
    datos = df.copy()
    y = datos["Iobs"].to_numpy(dtype=float)
    if len(y) < 3:
        datos["Linea_Base"] = np.zeros_like(y)
        datos["Iobs_Limpia"] = np.maximum(y, 0)
        return datos

    ventana = int(max(5, ventana))
    if ventana % 2 == 0:
        ventana += 1
    mitad = ventana // 2
    base = np.empty_like(y)
    for i in range(len(y)):
        ini = max(0, i - mitad)
        fin = min(len(y), i + mitad + 1)
        base[i] = np.percentile(y[ini:fin], percentil)

    suavizado = int(max(3, suavizado))
    if suavizado % 2 == 0:
        suavizado += 1
    kernel = np.ones(suavizado) / suavizado
    base = np.convolve(np.pad(base, (suavizado // 2,), mode="edge"), kernel, mode="valid")
    base = np.minimum(base, y)

    datos["Linea_Base"] = base
    datos["Iobs_Limpia"] = np.maximum(y - base, 0)
    return datos


def IdentificarPicos(df, pct_altura=0.08, pct_prominencia=0.06, distancia_minima=0.8):
    """
    Detecta maximos locales usando altura, prominencia aproximada y separacion
    minima entre picos.
    """
    x = df["2Theta"].to_numpy(dtype=float)
    y = df["Iobs_Limpia"].to_numpy(dtype=float)
    if len(y) < 3 or np.max(y) <= 0:
        return pd.DataFrame(columns=["Angulo_2Theta", "Iobs_maxima"])

    altura_min = np.max(y) * pct_altura
    prominencia_min = np.max(y) * pct_prominencia
    candidatos = []
    for i in range(1, len(y) - 1):
        if y[i] <= altura_min or not (y[i] >= y[i - 1] and y[i] >= y[i + 1]):
            continue
        izquierda = np.min(y[max(0, i - 20):i + 1])
        derecha = np.min(y[i:min(len(y), i + 21)])
        prominencia = y[i] - max(izquierda, derecha)
        if prominencia >= prominencia_min:
            candidatos.append((i, y[i]))

    candidatos.sort(key=lambda item: item[1], reverse=True)
    elegidos = []
    for indice, _ in candidatos:
        if all(abs(x[indice] - x[otro]) >= distancia_minima for otro in elegidos):
            elegidos.append(indice)

    elegidos.sort(key=lambda idx: x[idx])
    return pd.DataFrame({
        "Angulo_2Theta": x[elegidos],
        "Iobs_maxima": y[elegidos],
    })


def CompararConReferencia(df_resultados, df_referencia, tolerancia=0.3):
    """
    Asigna el indice de Miller mas cercano a cada pico experimental.
    """
    resultados = df_resultados.copy()
    etiquetas = []
    diferencias = []
    ref_angulos = df_referencia["2Theta"].to_numpy(dtype=float)

    for angulo in resultados["Angulo_2Theta"]:
        idx = int(np.argmin(np.abs(ref_angulos - angulo)))
        diferencia = abs(ref_angulos[idx] - angulo)
        diferencias.append(round(diferencia, 4))
        if diferencia <= tolerancia:
            etiquetas.append(str(df_referencia.iloc[idx]["Indice"]))
        else:
            etiquetas.append("* Impureza")

    resultados.insert(0, "Plano_hkl", etiquetas)
    resultados["Delta_2Theta"] = diferencias
    return resultados


def CalcularFWHM(df, df_picos, ventana_grados=0.6):
    """
    Calcula FWHM por interpolacion lineal a media altura sobre la senal limpia.
    """
    x = df["2Theta"].to_numpy(dtype=float)
    y = df["Iobs_Limpia"].to_numpy(dtype=float)
    filas = []

    for _, pico in df_picos.iterrows():
        centro = float(pico["Angulo_2Theta"])
        idx = int(np.argmin(np.abs(x - centro)))
        altura = y[idx]
        if altura <= 0:
            continue
        media_altura = altura / 2

        izq = idx
        while izq > 0 and x[idx] - x[izq] <= ventana_grados and y[izq] > media_altura:
            izq -= 1
        der = idx
        while der < len(y) - 1 and x[der] - x[idx] <= ventana_grados and y[der] > media_altura:
            der += 1
        if izq == 0 or der == len(y) - 1 or izq == idx or der == idx:
            continue

        x_izq = np.interp(media_altura, [y[izq], y[izq + 1]], [x[izq], x[izq + 1]])
        x_der = np.interp(media_altura, [y[der], y[der - 1]], [x[der], x[der - 1]])
        fwhm = abs(x_der - x_izq)
        if fwhm > 0:
            filas.append({
                "Angulo_2Theta": centro,
                "Iobs_maxima": altura,
                "FWHM_grados": fwhm,
            })

    return pd.DataFrame(filas)


def AplicarScherrer(df_resultados, longitud_onda=0.15406, K=0.9):
    """
    Aplica D = K*lambda / (beta*cos(theta)).
    longitud_onda debe estar en nm para obtener D en nm.
    """
    resultados = df_resultados.copy()
    theta = np.radians(resultados["Angulo_2Theta"].to_numpy(dtype=float) / 2)
    beta = np.radians(resultados["FWHM_grados"].to_numpy(dtype=float))
    with np.errstate(divide="ignore", invalid="ignore"):
        d_nm = (K * longitud_onda) / (beta * np.cos(theta))
    resultados["Tamaño_Cristal_nm"] = np.round(d_nm, 2)
    return resultados.replace([np.inf, -np.inf], np.nan)


def GenerarGrafica(df_curva, df_resultados, guardar_en=None, mostrar=True):
    """
    Grafica intensidad vs angulo para UN solo ensayo, marca picos y etiqueta
    indices hkl. Se mantiene igual que antes para no romper el flujo actual
    de un solo archivo experimental.
    """
    y_col = "Iobs_Limpia" if "Iobs_Limpia" in df_curva.columns else "Iobs"
    fig, ax = plt.subplots(figsize=(13, 7))
    fig.patch.set_facecolor("#f7f8fb")
    ax.set_facecolor("#ffffff")
    ax.plot(df_curva["2Theta"], df_curva[y_col], color="#202124", linewidth=1.15, label="Intensidad corregida")
    if "Linea_Base" in df_curva.columns and y_col != "Iobs":
        ax.plot(df_curva["2Theta"], df_curva["Linea_Base"], color="#9aa0a6",
                linewidth=0.9, alpha=0.65, label="Linea base")

    ymax = max(float(df_curva[y_col].max()), 1.0)
    puntos_x = []
    puntos_y = []
    for n, (_, fila) in enumerate(df_resultados.iterrows()):
        angulo = float(fila["Angulo_2Theta"])
        etiqueta = str(fila.get("Plano_hkl", "Pico"))
        color = "#b3261e" if "Impureza" in etiqueta else "#1a73e8"
        idx = int(np.argmin(np.abs(df_curva["2Theta"].to_numpy(dtype=float) - angulo)))
        intensidad_pico = float(df_curva[y_col].iloc[idx])
        puntos_x.append(angulo)
        puntos_y.append(intensidad_pico)
        y_texto = min(intensidad_pico + ymax * (0.055 + 0.045 * (n % 3)), ymax * 1.24)
        ax.axvline(angulo, color=color, linestyle="--", linewidth=0.75, alpha=0.45)
        ax.text(
            angulo, y_texto, etiqueta, rotation=90, ha="center", va="bottom",
            fontsize=8.5, color=color,
            bbox=dict(boxstyle="round,pad=0.22", facecolor="#ffffff",
                      edgecolor=color, linewidth=0.7, alpha=0.96),
        )
    if puntos_x:
        ax.scatter(puntos_x, puntos_y, s=32, color="#1a73e8", edgecolor="white",
                   linewidth=0.8, zorder=4, label="Picos detectados")

    if "Tamaño_Cristal_nm" in df_resultados.columns and not df_resultados.empty:
        promedio = df_resultados["Tamaño_Cristal_nm"].dropna().mean()
        texto = f"Picos: {len(df_resultados)}\nD promedio: {promedio:.2f} nm"
        ax.text(0.015, 0.96, texto, transform=ax.transAxes, ha="left", va="top",
                fontsize=10, color="#202124",
                bbox=dict(boxstyle="round,pad=0.45", facecolor="#eef3ff",
                          edgecolor="#c7d7fe", linewidth=0.8))

    ax.set_xlabel("2θ (°)")
    ax.set_ylabel("Intensidad (u.a.)")
    ax.set_title("Analisis DRX - Picos indexados", fontsize=15, pad=12)
    ax.set_ylim(0, ymax * 1.42)
    ax.grid(True, color="#dfe3ea", linestyle="-", linewidth=0.6, alpha=0.75)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#c7cdd8")
    ax.spines["bottom"].set_color("#c7cdd8")
    ax.legend(loc="upper right", frameon=True, facecolor="white", edgecolor="#dfe3ea")
    fig.tight_layout()

    if guardar_en:
        fig.savefig(guardar_en, dpi=200)
    if mostrar:
        plt.show()
    return fig


_MARCADORES_IMPUREZA = {
    "mno2": ("v", "#d32f2f", "Impureza MnO$_2$"),
    "li2co3": ("D", "#f28c00", "Impureza Li$_2$CO$_3$"),
    "li2nio2": ("o", "#202124", "Impureza Li$_2$NiO$_2$"),
    "ni2o3": ("s", "#1a237e", "Impureza Ni$_2$O$_3$"),
}


def _clave_impureza(etiqueta):
    normalizada = etiqueta.lower().replace(" ", "").replace("₂", "2").replace("*", "")
    for clave in _MARCADORES_IMPUREZA:
        if clave in normalizada:
            return clave
    return None


def GenerarGraficaComparativa(df_curva_1, df_resultados_1, df_curva_2,
                               etiqueta_1="Ensayo 1", etiqueta_2="Ensayo 2",
                               df_referencia=None, nombre_referencia="Patron",
                               guardar_en=None, mostrar=True):
    """
    Grafica DOS ensayos experimentales en la misma imagen, uno arriba
    (etiqueta_1, en negro) y otro abajo (etiqueta_2, en gris), cada uno
    normalizado a su propio maximo y separados por un offset vertical fijo.
    Los picos del ensayo 1 (el que tiene indexacion / Plano_hkl) se marcan
    con lineas punteadas verticales que atraviesan ambas curvas, para poder
    comparar visualmente si la misma señal aparece en el segundo ensayo.

    Si se pasa df_referencia (salida de CargarReferencia), se dibuja una
    fila de marcas en la parte inferior con las posiciones teoricas, similar
    a la fila "Patron" de un difractograma de PDF.
    """
    y1_col = "Iobs_Limpia" if "Iobs_Limpia" in df_curva_1.columns else "Iobs"
    y2_col = "Iobs_Limpia" if "Iobs_Limpia" in df_curva_2.columns else "Iobs"

    x1 = df_curva_1["2Theta"].to_numpy(dtype=float)
    x2 = df_curva_2["2Theta"].to_numpy(dtype=float)
    y1 = df_curva_1[y1_col].to_numpy(dtype=float)
    y2 = df_curva_2[y2_col].to_numpy(dtype=float)

    y1n = y1 / max(y1.max(), 1e-9)
    y2n = y2 / max(y2.max(), 1e-9)
    offset = 1.15

    fig, ax = plt.subplots(figsize=(14, 7.5))
    fig.patch.set_facecolor("#f7f8fb")
    ax.set_facecolor("#ffffff")

    ax.plot(x2, y2n, color="#9aa0a6", linewidth=1.0, label=etiqueta_2)
    ax.plot(x1, y1n + offset, color="#202124", linewidth=1.15, label=etiqueta_1)

    etiquetas_leyenda_usadas = set()

    if df_resultados_1 is not None and not df_resultados_1.empty:
        for n, (_, fila) in enumerate(df_resultados_1.iterrows()):
            angulo = float(fila["Angulo_2Theta"])
            etiqueta = str(fila.get("Plano_hkl", "Pico"))
            ax.axvline(angulo, color="#c7cdd8", linestyle=":", linewidth=0.8, alpha=0.85, zorder=1)

            idx1 = int(np.argmin(np.abs(x1 - angulo)))
            altura_pico = y1n[idx1] + offset

            es_impureza = "impureza" in etiqueta.lower() or "*" in etiqueta
            if es_impureza:
                clave = _clave_impureza(etiqueta)
                marcador, color, nombre_leyenda = _MARCADORES_IMPUREZA.get(
                    clave, ("x", "#5f6368", "Impureza")
                )
                ax.scatter(
                    [angulo], [altura_pico], marker=marcador, s=60, color=color,
                    edgecolor="white", linewidth=0.6, zorder=5,
                    label=nombre_leyenda if nombre_leyenda not in etiquetas_leyenda_usadas else None,
                )
                etiquetas_leyenda_usadas.add(nombre_leyenda)
            else:
                y_texto = min(altura_pico + 0.08 + 0.04 * (n % 3), offset + 1.35)
                ax.text(
                    angulo, y_texto, etiqueta, rotation=90, ha="center", va="bottom",
                    fontsize=8, color="#1a73e8",
                    bbox=dict(boxstyle="round,pad=0.2", facecolor="white",
                              edgecolor="#1a73e8", linewidth=0.6, alpha=0.95),
                )

    if df_referencia is not None and not df_referencia.empty:
        y_base = -0.35
        x_min = min(x1.min(), x2.min())
        x_max = max(x1.max(), x2.max())
        ax.plot([x_min, x_max], [y_base, y_base], color="#202124", linewidth=1.0)
        ax.vlines(df_referencia["2Theta"], y_base - 0.03, y_base + 0.03, color="#202124", linewidth=1.0)
        ax.text(x_min, y_base, f"  {nombre_referencia}", va="center", fontsize=9)
        ax.set_ylim(y_base - 0.15, offset + 1.45)
    else:
        ax.set_ylim(-0.1, offset + 1.45)

    ax.set_xlabel("2θ (°)")
    ax.set_ylabel("Intensidad [u.a.]")
    ax.set_title("Comparación de ensayos DRX", fontsize=15, pad=12)
    ax.set_yticks([])
    ax.grid(True, axis="x", color="#dfe3ea", linestyle="-", linewidth=0.6, alpha=0.6)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.legend(loc="upper left", frameon=True, facecolor="white", edgecolor="#dfe3ea")
    fig.tight_layout()

    if guardar_en:
        fig.savefig(guardar_en, dpi=200)
    if mostrar:
        plt.show()
    return fig


def GenerarReporte(df_resultados, ruta_salida=None):
    """
    Exporta la tabla final con plano, angulo, FWHM y tamano de cristalito.
    """
    columnas = [c for c in [
        "Plano_hkl", "Angulo_2Theta", "FWHM_grados", "Tamaño_Cristal_nm", "Delta_2Theta"
    ] if c in df_resultados.columns]
    reporte = df_resultados[columnas].copy()
    reporte = reporte.rename(columns={
        "Plano_hkl": "Plano hkl",
        "Angulo_2Theta": "Angulo 2Theta (grados)",
        "FWHM_grados": "FWHM (grados)",
        "Tamaño_Cristal_nm": "Tamano cristalito D (nm)",
        "Delta_2Theta": "Diferencia referencia (grados)",
    })
    for columna in reporte.select_dtypes(include=[np.number]).columns:
        reporte[columna] = reporte[columna].round(4)
    if ruta_salida:
        extension = os.path.splitext(ruta_salida)[1].lower()
        if extension in (".html", ".htm"):
            _exportar_reporte_html(reporte, ruta_salida)
        elif extension == ".xlsx":
            _exportar_reporte_xlsx(reporte, ruta_salida)
        else:
            reporte.to_csv(ruta_salida, index=False, encoding="utf-8-sig", sep=";", decimal=",")
    return reporte


def _exportar_reporte_html(reporte, ruta_salida):
    promedio = reporte["Tamano cristalito D (nm)"].mean() if "Tamano cristalito D (nm)" in reporte else np.nan
    filas = []
    for _, fila in reporte.iterrows():
        es_impureza = "Impureza" in str(fila.get("Plano hkl", ""))
        clase = "impureza" if es_impureza else "fase"
        celdas = "".join(f"<td>{valor}</td>" for valor in fila)
        filas.append(f"<tr class='{clase}'>{celdas}</tr>")
    encabezados = "".join(f"<th>{col}</th>" for col in reporte.columns)
    html = f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>Reporte DRX Scherrer</title>
  <style>
    body {{ font-family: Segoe UI, Arial, sans-serif; margin: 36px; color: #202124; background: #f7f8fb; }}
    .header {{ border-left: 6px solid #1a73e8; padding-left: 18px; margin-bottom: 22px; }}
    h1 {{ margin: 0 0 6px; font-size: 28px; }}
    .sub {{ color: #5f6368; font-size: 14px; }}
    .cards {{ display: flex; gap: 14px; margin: 22px 0; }}
    .card {{ background: white; border: 1px solid #dfe3ea; border-radius: 8px; padding: 14px 18px; min-width: 160px; }}
    .label {{ color: #5f6368; font-size: 12px; text-transform: uppercase; letter-spacing: .04em; }}
    .value {{ font-size: 22px; font-weight: 700; margin-top: 4px; }}
    table {{ border-collapse: collapse; width: 100%; background: white; border: 1px solid #dfe3ea; }}
    th {{ background: #1f2937; color: white; text-align: center; padding: 10px; font-size: 13px; }}
    td {{ padding: 9px 10px; border-bottom: 1px solid #eef0f4; text-align: center; font-size: 13px; }}
    tr:nth-child(even) td {{ background: #fbfcff; }}
    tr.impureza td:first-child {{ color: #b3261e; font-weight: 700; }}
    tr.fase td:first-child {{ color: #1a73e8; font-weight: 700; }}
    .nota {{ margin-top: 18px; color: #5f6368; font-size: 12px; }}
  </style>
</head>
<body>
  <div class="header">
    <h1>Reporte cristalografico DRX</h1>
    <div class="sub">Resultados de indexacion, FWHM y tamano de cristalito por Scherrer</div>
  </div>
  <div class="cards">
    <div class="card"><div class="label">Picos procesados</div><div class="value">{len(reporte)}</div></div>
    <div class="card"><div class="label">D promedio</div><div class="value">{promedio:.2f} nm</div></div>
  </div>
  <table>
    <thead><tr>{encabezados}</tr></thead>
    <tbody>{''.join(filas)}</tbody>
  </table>
  <div class="nota">Los picos marcados como impureza superan la tolerancia configurada contra la referencia.</div>
</body>
</html>"""
    with open(ruta_salida, "w", encoding="utf-8") as archivo:
        archivo.write(html)


def _exportar_reporte_xlsx(reporte, ruta_salida):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Para exportar XLSX con formato instala openpyxl: pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte DRX"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(reporte.columns))
    ws.cell(1, 1).value = "Reporte cristalografico DRX"
    ws.cell(1, 1).font = Font(size=16, bold=True, color="1F2937")
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DFE3EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, nombre in enumerate(reporte.columns, start=1):
        cell = ws.cell(3, col_idx, nombre)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, (_, fila) in enumerate(reporte.iterrows(), start=4):
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row_idx, col_idx, valor)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            if col_idx == 1:
                cell.font = Font(bold=True, color="B3261E" if "Impureza" in str(valor) else "1A73E8")

    for col_idx, nombre in enumerate(reporte.columns, start=1):
        largo = max(len(str(nombre)), *(len(str(v)) for v in reporte[nombre].astype(str))) + 3
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(largo, 14), 34)
    ws.freeze_panes = "A4"
    wb.save(ruta_salida)


def AnalizarDRX(ruta_experimental, ruta_referencia, tolerancia=0.3,
                longitud_onda=0.15406, K=0.9):
    df = CargarDatos(ruta_experimental)
    referencia = CargarReferencia(ruta_referencia)
    df_limpio = RemoverBackground(df)
    picos = IdentificarPicos(df_limpio)
    minimo_ref = float(referencia["2Theta"].min()) - max(float(tolerancia), 0.0)
    maximo_ref = float(referencia["2Theta"].max()) + max(float(tolerancia), 0.0)
    picos = picos[(picos["Angulo_2Theta"] >= minimo_ref) & (picos["Angulo_2Theta"] <= maximo_ref)]
    resultados = CalcularFWHM(df_limpio, picos)
    if resultados.empty:
        raise ValueError("No se pudieron calcular FWHM. Revisa los parametros o el archivo.")
    resultados = AplicarScherrer(resultados, longitud_onda=longitud_onda, K=K)
    resultados = CompararConReferencia(resultados, referencia, tolerancia=tolerancia)
    return df_limpio, resultados, referencia


class InterfazDRX(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Analizador DRX - Scherrer")
        self.geometry("1220x780")
        self.minsize(1000, 660)

        self.ruta_exp = tk.StringVar()
        self.ruta_exp2 = tk.StringVar()
        self.ruta_ref = tk.StringVar()
        self.tolerancia = tk.DoubleVar(value=0.3)
        self.lambda_nm = tk.DoubleVar(value=0.15406)
        self.k = tk.DoubleVar(value=0.9)
        self.df_limpio = None
        self.df_limpio2 = None
        self.df_resultados = None
        self.referencia = None
        self.canvas = None

        self._crear_estilos()
        self._crear_widgets()

    def _crear_estilos(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TFrame", background="#f6f7f9")
        style.configure("TLabel", background="#f6f7f9", font=("Segoe UI", 10))
        style.configure("Title.TLabel", font=("Segoe UI", 18, "bold"))
        style.configure("TButton", font=("Segoe UI", 10), padding=7)
        style.configure("Treeview", rowheight=26, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))

    def _crear_widgets(self):
        contenedor = ttk.Frame(self, padding=14)
        contenedor.pack(fill="both", expand=True)

        encabezado = ttk.Frame(contenedor)
        encabezado.pack(fill="x")
        ttk.Label(encabezado, text="Analizador DRX", style="Title.TLabel").pack(side="left")
        ttk.Label(encabezado, text="Carga datos, indexa picos y calcula Scherrer").pack(side="left", padx=14)

        controles = ttk.Frame(contenedor, padding=(0, 12, 0, 10))
        controles.pack(fill="x")
        self._fila_archivo(controles, "Experimental 1", self.ruta_exp, self._seleccionar_exp, 0)
        self._fila_archivo(controles, "Experimental 2 (opcional)", self.ruta_exp2, self._seleccionar_exp2, 1)
        self._fila_archivo(controles, "Referencia", self.ruta_ref, self._seleccionar_ref, 2)

        parametros = ttk.Frame(controles)
        parametros.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(8, 0))
        for texto, variable, ancho in [
            ("Tolerancia 2θ", self.tolerancia, 8),
            ("λ Cu Kα (nm)", self.lambda_nm, 10),
            ("K", self.k, 6),
        ]:
            ttk.Label(parametros, text=texto).pack(side="left", padx=(0, 5))
            ttk.Entry(parametros, textvariable=variable, width=ancho).pack(side="left", padx=(0, 18))
        ttk.Button(parametros, text="Analizar", command=self._analizar).pack(side="left", padx=(6, 0))
        ttk.Button(parametros, text="Exportar reporte", command=self._exportar).pack(side="left", padx=6)
        ttk.Button(parametros, text="Guardar grafica", command=self._guardar_grafica).pack(side="left", padx=(0, 6))
        ttk.Button(parametros, text="Ver info CIF", command=self._cargar_cif).pack(side="left")
        controles.columnconfigure(1, weight=1)

        cuerpo = ttk.PanedWindow(contenedor, orient="horizontal")
        cuerpo.pack(fill="both", expand=True)

        tabla_frame = ttk.Frame(cuerpo)
        self.tabla = ttk.Treeview(tabla_frame, columns=("plano", "angulo", "fwhm", "tamano", "delta"), show="headings")
        for col, titulo, ancho in [
            ("plano", "Plano hkl", 90),
            ("angulo", "2θ", 90),
            ("fwhm", "FWHM", 90),
            ("tamano", "D (nm)", 90),
            ("delta", "Δ ref.", 90),
        ]:
            self.tabla.heading(col, text=titulo)
            self.tabla.column(col, width=ancho, anchor="center")
        scroll = ttk.Scrollbar(tabla_frame, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=scroll.set)
        self.tabla.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        self.grafica_frame = ttk.Frame(cuerpo)
        cuerpo.add(tabla_frame, weight=1)
        cuerpo.add(self.grafica_frame, weight=3)

        self.estado = ttk.Label(contenedor, text="Listo para cargar archivos.")
        self.estado.pack(fill="x", pady=(8, 0))

    def _fila_archivo(self, parent, etiqueta, variable, comando, fila):
        ttk.Label(parent, text=etiqueta).grid(row=fila, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(parent, textvariable=variable).grid(row=fila, column=1, sticky="ew", pady=3)
        ttk.Button(parent, text="Buscar", command=comando).grid(row=fila, column=2, padx=(8, 0), pady=3)

    def _seleccionar_exp(self):
        ruta = filedialog.askopenfilename(title="Archivo experimental 1", filetypes=[("Datos", "*.txt *.xy *.dat *.csv"), ("Todos", "*.*")])
        if ruta:
            self.ruta_exp.set(ruta)

    def _seleccionar_exp2(self):
        ruta = filedialog.askopenfilename(title="Archivo experimental 2 (opcional)", filetypes=[("Datos", "*.txt *.xy *.dat *.csv"), ("Todos", "*.*")])
        if ruta:
            self.ruta_exp2.set(ruta)

    def _seleccionar_ref(self):
        ruta = filedialog.askopenfilename(title="Archivo de referencia", filetypes=[("Datos", "*.txt *.dat *.csv"), ("Todos", "*.*")])
        if ruta:
            self.ruta_ref.set(ruta)

    def _analizar(self):
        try:
            if not self.ruta_exp.get() or not self.ruta_ref.get():
                raise ValueError("Selecciona al menos el archivo experimental 1 y el de referencia.")
            self.estado.config(text="Analizando datos...")
            self.update_idletasks()
            self.df_limpio, self.df_resultados, self.referencia = AnalizarDRX(
                self.ruta_exp.get(),
                self.ruta_ref.get(),
                tolerancia=float(self.tolerancia.get()),
                longitud_onda=float(self.lambda_nm.get()),
                K=float(self.k.get()),
            )

            self.df_limpio2 = None
            if self.ruta_exp2.get():
                df2 = CargarDatos(self.ruta_exp2.get())
                self.df_limpio2 = RemoverBackground(df2)

            self._actualizar_tabla()
            self._actualizar_grafica()
            mensaje = f"Analisis completo: {len(self.df_resultados)} picos procesados."
            if self.df_limpio2 is not None:
                mensaje += " Mostrando comparacion con el segundo ensayo."
            self.estado.config(text=mensaje)
        except Exception as exc:
            self.estado.config(text="Error durante el analisis.")
            messagebox.showerror("Error", str(exc))

    def _actualizar_tabla(self):
        for item in self.tabla.get_children():
            self.tabla.delete(item)
        for _, fila in self.df_resultados.iterrows():
            self.tabla.insert("", "end", values=(
                fila.get("Plano_hkl", ""),
                f"{fila.get('Angulo_2Theta', np.nan):.4f}",
                f"{fila.get('FWHM_grados', np.nan):.4f}",
                f"{fila.get('Tamaño_Cristal_nm', np.nan):.2f}",
                f"{fila.get('Delta_2Theta', np.nan):.4f}",
            ))

    def _generar_figura(self, guardar_en=None):
        if self.df_limpio2 is not None:
            return GenerarGraficaComparativa(
                self.df_limpio, self.df_resultados, self.df_limpio2,
                etiqueta_1=os.path.basename(self.ruta_exp.get()) or "Ensayo 1",
                etiqueta_2=os.path.basename(self.ruta_exp2.get()) or "Ensayo 2",
                df_referencia=self.referencia,
                guardar_en=guardar_en, mostrar=False,
            )
        return GenerarGrafica(self.df_limpio, self.df_resultados, guardar_en=guardar_en, mostrar=False)

    def _actualizar_grafica(self):
        for widget in self.grafica_frame.winfo_children():
            widget.destroy()
        fig = self._generar_figura()
        self.canvas = FigureCanvasTkAgg(fig, master=self.grafica_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)

    def _exportar(self):
        if self.df_resultados is None:
            messagebox.showinfo("Exportar", "Primero ejecuta el analisis.")
            return
        ruta = filedialog.asksaveasfilename(
            title="Guardar reporte",
            defaultextension=".html",
            initialfile="reporte_drx.html",
            filetypes=[
                ("Reporte HTML lindo", "*.html"),
                ("CSV compatible Excel", "*.csv"),
                ("Excel con formato", "*.xlsx"),
            ],
        )
        if ruta:
            try:
                GenerarReporte(self.df_resultados, ruta)
                self.estado.config(text=f"Reporte guardado en {os.path.basename(ruta)}.")
            except Exception as exc:
                messagebox.showerror("Exportar reporte", str(exc))

    def _guardar_grafica(self):
        if self.df_resultados is None or self.df_limpio is None:
            messagebox.showinfo("Guardar grafica", "Primero ejecuta el analisis.")
            return
        ruta = filedialog.asksaveasfilename(
            title="Guardar grafica",
            defaultextension=".png",
            initialfile="grafica_drx.png",
            filetypes=[("PNG", "*.png"), ("PDF", "*.pdf")],
        )
        if ruta:
            fig = self._generar_figura(guardar_en=ruta)
            plt.close(fig)
            self.estado.config(text=f"Grafica guardada en {os.path.basename(ruta)}.")

    def _cargar_cif(self):
        try:
            info = CargarCIF()
            if info is None:
                return
            celda = info["celda"]
            texto = (
                f"Archivo: {os.path.basename(info['ruta'])}\n\n"
                f"a = {celda['a']}\nb = {celda['b']}\nc = {celda['c']}\n"
                f"alpha = {celda['alpha']}\nbeta = {celda['beta']}\ngamma = {celda['gamma']}\n"
                f"Volumen = {celda['volumen']}\n\n"
                f"Grupo espacial: {info['grupo_espacial']}\n"
                f"Atomos leidos: {len(info['atomos'])}"
            )
            messagebox.showinfo("Informacion CIF", texto)
        except Exception as exc:
            messagebox.showerror("Leer CIF", str(exc))


def iniciar_interfaz():
    app = InterfazDRX()
    app.mainloop()


if __name__ == "__main__":
    iniciar_interfaz()
